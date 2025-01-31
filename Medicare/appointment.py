from flask import Flask, request, jsonify
import openpyxl
import os

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = 'PatientDetails.xlsx'

# Function to add patient details to the Excel file
def add_patient_to_excel(name, age, contact):
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Age', 'Contact'])  # Header
    else:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

    sheet.append([name, age, contact])  # Add patient details
    workbook.save(EXCEL_FILE)

# Route to handle form submission
@app.route('/submit', methods=['POST'])
def submit():
    data = request.json
    name = data.get('name')
    age = data.get('age')
    contact = data.get('contact')

    add_patient_to_excel(name, age, contact)

    # Count total patients
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    total_patients = sheet.max_row - 1  # Exclude header

    return jsonify({'totalPatients': total_patients})

if __name__ == '__main__':
    app.run(debug=True)
